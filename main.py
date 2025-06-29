import sys
import os
import json
from datetime import datetime
from flask import Flask, render_template, redirect, url_for, request, flash, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from flask_dance.contrib.google import make_google_blueprint, google
from flask_dance.consumer import oauth_authorized
from flask_session import Session
from werkzeug.middleware.proxy_fix import ProxyFix
import openpyxl
from pymongo import MongoClient
from bson.objectid import ObjectId

# Load environment variables from .env file if it exists
def load_env_file():
    env_file = '.env'
    if os.path.exists(env_file):
        with open(env_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key.strip()] = value.strip()

# Load .env file
load_env_file()

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
app.secret_key = 'sanchari_mentors_secret_key'

# Enhanced session configuration to prevent OAuth state mismatch
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_COOKIE_SECURE'] = False  # Set to True in production
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hour
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
app.config['SESSION_FILE_DIR'] = 'flask_session'
app.config['SESSION_FILE_THRESHOLD'] = 500
app.config['SESSION_FILE_MODE'] = 0o600
app.config['SESSION_KEY_PREFIX'] = 'sanchari_'
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_COOKIE_NAME'] = 'sanchari_session'

# Create session directory if it doesn't exist
os.makedirs('flask_session', exist_ok=True)

Session(app)

# MongoDB connection
MONGO_URI = os.environ.get('MONGO_URI', 'mongodb://localhost:27017/sanchari_mentors')
mongo_client = MongoClient(MONGO_URI)
mongo_db = mongo_client.get_default_database()

def get_db():
    return mongo_db

# Utility functions for MongoDB (replacement for load_json_data/save_json_data)
def get_collection_data(collection_name, key_field=None):
    db = get_db()
    collection = db[collection_name]
    if key_field:
        # Return as dict keyed by key_field
        docs = list(collection.find())
        return {doc[key_field]: doc for doc in docs}
    else:
        return list(collection.find())

def set_collection_data(collection_name, data, key_field=None):
    db = get_db()
    collection = db[collection_name]
    collection.delete_many({})
    if isinstance(data, dict):
        # If data is a dict keyed by key_field
        docs = list(data.values())
        collection.insert_many(docs)
    elif isinstance(data, list):
        collection.insert_many(data)
    else:
        raise ValueError('Data must be a list or dict')

# Google OAuth setup
# To get these credentials:
# 1. Go to https://console.cloud.google.com/
# 2. Create a new project or select existing one
# 3. Enable Google+ API
# 4. Go to Credentials -> Create Credentials -> OAuth 2.0 Client IDs
# 5. Set authorized redirect URIs to: http://localhost:8000/login/google/authorized
# 6. Copy Client ID and Client Secret below

GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID')
GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET')

# For local development only - remove in production
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
# Suppress OAuth scope warnings
os.environ['OAUTHLIB_RELAX_TOKEN_SCOPE'] = '1'
# Fix OAuth state mismatch issues
os.environ['OAUTHLIB_RELAX_TOKEN_SCOPE'] = '1'
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'

# Check if Google credentials are properly set
if GOOGLE_CLIENT_ID == 'YOUR_GOOGLE_CLIENT_ID' or GOOGLE_CLIENT_SECRET == 'YOUR_GOOGLE_CLIENT_SECRET':
    print("⚠️  WARNING: Google OAuth credentials not configured!")
    print("Please set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET environment variables")
    print("or update the values in main.py")
    print("Google login will not work until credentials are configured.")

google_bp = make_google_blueprint(
    client_id=GOOGLE_CLIENT_ID,
    client_secret=GOOGLE_CLIENT_SECRET,
    scope=[
        "openid",
        "https://www.googleapis.com/auth/userinfo.email",
        "https://www.googleapis.com/auth/userinfo.profile"
    ],
    redirect_url="http://localhost:8000/login/google/authorized",
    reprompt_consent=True,
    reprompt_select_account=True,
    storage=None  # Use Flask session storage
)
app.register_blueprint(google_bp, url_prefix="/login")

# Signal handler for successful Google OAuth login
@oauth_authorized.connect_via(google_bp)
def google_logged_in(blueprint, token):
    if not token:
        flash("Failed to log in with Google.", "danger")
        return redirect(url_for('login'))

    try:
        resp = blueprint.session.get("/oauth2/v2/userinfo")
        if not resp.ok:
            msg = "Failed to fetch user info from Google."
            flash(msg, "danger")
            return redirect(url_for('login'))

        user_info = resp.json()
        email = user_info["email"]

        users_data = get_collection_data('users')
        user = users_data.get(email)

        # Auto-create user if not present
        if not user:
            user = {
                "name": user_info.get("name", email.split('@')[0]),
                "email": email,
                "role": "student",
                "interests": [],
                "profile_pic": user_info.get("picture", ""),
                "grade": None,
                "interested_subjects": []
            }
            users_data[email] = user
            set_collection_data('users', users_data)
            flash("Welcome! Your account has been created.", "success")

        # Set session data
        session['user_email'] = email
        session['user_role'] = user['role']
        session['user_name'] = user['name']
        session.permanent = True

        flash(f'Successfully logged in as {user["name"]}!', 'success')

        # Redirect to the correct dashboard
        if user['role'] == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif user['role'] == 'trainer':
            return redirect(url_for('trainer_dashboard'))
        else:
            if user['role'] == 'student' and (not user.get('grade') or not user.get('interested_subjects')):
                return redirect(url_for('complete_student_profile'))
            return redirect(url_for('student_dashboard'))

    except Exception as e:
        print(f"Error in google_logged_in signal handler: {e}")
        flash('An unexpected error occurred during Google login.', 'danger')
        return redirect(url_for('login'))

# Home page route
@app.route('/')
def index():
    return render_template('index.html')

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        users_data = get_collection_data('users')
        user = users_data.get(email)
        
        if user and user.get('password') == password:
            session['user_email'] = email
            session['user_role'] = user['role']
            session['user_name'] = user['name']
            flash(f'Welcome back, {session["user_name"]}!', 'success')
            
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif user['role'] == 'trainer':
                return redirect(url_for('trainer_dashboard'))
            else:
                return redirect(url_for('student_dashboard'))
        else:
            flash('Invalid email or password. Please try again.', 'danger')
            return redirect(url_for('login'))

    # If a user is already logged in, redirect them
    if 'user_email' in session:
        role = session.get('user_role')
        if role == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif role == 'trainer':
            return redirect(url_for('trainer_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    # Clear the user's session
    session.clear()
    
    # Invalidate the Google OAuth token if it exists
    if 'google_token' in session:
        del session['google_token']

    flash('You have been successfully logged out.', 'success')
    return redirect(url_for('index'))

# Student routes
@app.route('/student/dashboard')
def student_dashboard():
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    users_data = get_collection_data('users')
    user = users_data.get(session['user_email'], {})
    return render_template('student/dashboard.html', user=user)

@app.route('/student/classes')
def student_classes():
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    classes_data = get_collection_data('classes')
    return render_template('student/classes.html', classes=classes_data.get('classes', []))

@app.route('/student/resources')
def student_resources():
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    return render_template('student/resources.html')

@app.route('/student/doubts')
def student_doubts():
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    return render_template('student/doubts.html')

@app.route('/student/career')
def student_career():
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    return render_template('student/career.html')

@app.route('/student/skills')
def student_skills():
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    return render_template('student/skills.html')

@app.route('/student/achievements')
def student_achievements():
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    return render_template('student/achievements.html')

@app.route('/student/community')
def student_community():
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    return render_template('student/community.html')

@app.route('/student/interests', methods=['GET', 'POST'])
def student_interests():
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    users_data = get_collection_data('users')
    user = users_data.get(session['user_email'], {})
    user_interests = user.get('interests', [])
    
    if request.method == 'POST':
        interests = request.form.getlist('interests')
        user['interests'] = interests
        users_data[session['user_email']] = user
        set_collection_data('users', users_data)
        flash('Your interests have been updated successfully!')
        return redirect(url_for('student_interests'))
    
    return render_template('student/interests.html', user_interests=user_interests)

# Student resource detail route
@app.route('/student/resource/<int:resource_id>')
def student_resource_detail(resource_id):
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    # Mock resource data - you can create a resources.json file later
    mock_resources = {
        1: {
            "id": 1,
            "title": "Introduction to Python Programming",
            "type": "video",
            "description": "Learn the basics of Python programming language",
            "duration": "2 hours",
            "instructor": "Dr. Sarah Johnson",
            "url": "#",
            "thumbnail": "https://via.placeholder.com/300x200/4CAF50/FFFFFF?text=Python+Basics",
            "tags": ["python", "programming", "beginner"],
            "rating": 4.8,
            "views": 1250
        },
        2: {
            "id": 2,
            "title": "Web Development Fundamentals",
            "type": "document",
            "description": "Comprehensive guide to HTML, CSS, and JavaScript",
            "duration": "3 hours",
            "instructor": "Prof. Michael Chen",
            "url": "#",
            "thumbnail": "https://via.placeholder.com/300x200/2196F3/FFFFFF?text=Web+Dev",
            "tags": ["html", "css", "javascript", "web"],
            "rating": 4.6,
            "views": 890
        }
    }
    
    resource = mock_resources.get(resource_id)
    if not resource:
        flash('Resource not found.', 'danger')
        return redirect(url_for('student_resources'))
    
    return render_template('student/resource_detail.html', resource=resource)

# Student question detail route
@app.route('/student/question/<int:question_id>')
def student_question_detail(question_id):
    if 'user_role' not in session or session['user_role'] != 'student':
        flash('Please login as a student to access this page.')
        return redirect(url_for('login'))
    
    # Mock question data
    mock_questions = {
        1: {
            "id": 1,
            "title": "How to implement authentication in Flask?",
            "content": "I'm building a web application using Flask and need to implement user authentication. Can someone help me with the best practices for secure authentication?",
            "author": "student@example.com",
            "date": "2024-06-20",
            "category": "Programming",
            "status": "answered",
            "answers": [
                {
                    "id": 1,
                    "content": "You can use Flask-Login for authentication. Here's a basic example...",
                    "author": "trainer@example.com",
                    "date": "2024-06-21",
                    "is_accepted": True
                }
            ],
            "tags": ["flask", "authentication", "python"]
        }
    }
    
    question = mock_questions.get(question_id)
    if not question:
        flash('Question not found.', 'danger')
        return redirect(url_for('student_doubts'))
    
    return render_template('student/question_detail.html', question=question)

# Admin routes
@app.route('/admin/dashboard')
def admin_dashboard():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))

    users_data = get_collection_data('users')
    classes_data = get_collection_data('classes')
    sessions_data = get_collection_data('sessions')

    total_students = len([u for u in users_data.values() if u.get('role') == 'student'])
    total_trainers = len([u for u in users_data.values() if u.get('role') == 'trainer'])
    total_classes = len(classes_data.get('classes', []))
    total_sessions = len(sessions_data.get('sessions', []))
    attendance_records = [s for s in sessions_data.get('sessions', []) if s.get('attendance_status') == 'present']
    attendance_rate = 0
    if total_sessions > 0:
        attendance_rate = int(100 * len(attendance_records) / total_sessions)

    return render_template(
        'admin/dashboard.html',
        total_students=total_students,
        total_trainers=total_trainers,
        total_classes=total_classes,
        total_sessions=total_sessions,
        attendance_rate=attendance_rate
    )

@app.route('/admin/users')
def admin_users():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))
    
    users_data = get_collection_data('users')
    return render_template('admin/users.html', users=users_data)

@app.route('/admin/create-user', methods=['GET', 'POST'])
def admin_create_user():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        users_data = get_collection_data('users')
        
        email = request.form.get('email')
        name = request.form.get('name')
        role = request.form.get('role')
        password = request.form.get('password')
        
        if email in users_data:
            flash('User with this email already exists.')
        else:
            users_data[email] = {
                'name': name,
                'email': email,
                'role': role,
                'password': generate_password_hash(password) if password else '',
                'interests': [],
                'created_at': datetime.now().isoformat()
            }
            set_collection_data('users', users_data)
            flash('User created successfully!')
            return redirect(url_for('admin_users'))
    
    return render_template('admin/create_user.html')

@app.route('/admin/edit-user/<email>', methods=['GET', 'POST'])
def admin_edit_user(email):
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))
    
    users_data = get_collection_data('users')
    user = users_data.get(email)
    
    if not user:
        flash('User not found.')
        return redirect(url_for('admin_users'))
    
    if request.method == 'POST':
        user['name'] = request.form.get('name')
        user['role'] = request.form.get('role')
        password = request.form.get('password')
        if password:
            user['password'] = generate_password_hash(password)
        
        users_data[email] = user
        set_collection_data('users', users_data)
        flash('User updated successfully!')
        return redirect(url_for('admin_users'))
    
    return render_template('admin/edit_user.html', user=user)

@app.route('/admin/trainers')
def admin_trainers():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))
    
    return render_template('admin/trainers.html')

@app.route('/admin/classes')
def admin_classes():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))
    
    classes_data = get_collection_data('classes')
    return render_template('admin/classes.html', classes=classes_data.get('classes', []))

@app.route('/admin/sessions')
def admin_sessions():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))
    
    sessions_data = get_collection_data('sessions')
    return render_template('admin/sessions.html', sessions=sessions_data.get('sessions', []))

@app.route('/admin/content')
def admin_content():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))
    
    return render_template('admin/content.html')

@app.route('/admin/reports')
def admin_reports():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))
    
    return render_template('admin/reports.html')

@app.route('/admin/communications')
def admin_communications():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('Please login as admin to access this page.')
        return redirect(url_for('login'))
    
    return render_template('admin/communications.html')

@app.route('/admin/download-career-assessments')
def download_career_assessments():
    if 'user_role' not in session or session['user_role'] != 'admin':
        flash('You do not have permission to access this resource.', 'danger')
        return redirect(url_for('login'))
    file_path = 'career_assessments.xlsx'
    import os
    if not os.path.exists(file_path):
        flash('No career assessments have been submitted yet.', 'warning')
        return redirect(url_for('admin_dashboard'))
    return send_file(file_path, as_attachment=True)

# Trainer routes
@app.route('/trainer/dashboard')
def trainer_dashboard():
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    return render_template('trainer/dashboard.html')

@app.route('/trainer/sessions')
def trainer_sessions():
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    return render_template('trainer/sessions.html')

@app.route('/trainer/students')
def trainer_students():
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    return render_template('trainer/students.html')

@app.route('/trainer/resources')
def trainer_resources():
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    return render_template('trainer/resources.html')

@app.route('/trainer/communications')
def trainer_communications():
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    return render_template('trainer/communications.html')

@app.route('/trainer/training')
def trainer_training():
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    return render_template('trainer/training.html')

@app.route('/trainer/feedback')
def trainer_feedback():
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    return render_template('trainer/feedback.html')

@app.route('/trainer/feedback/<int:feedback_id>')
def trainer_feedback_detail(feedback_id):
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    # Mock feedback data - you can create a feedback.json file later
    feedback_data = {
        'id': feedback_id,
        'student_name': 'John Doe',
        'student_email': 'john.doe@example.com',
        'class_title': 'Introduction to AI',
        'rating': 4,
        'comment': 'Great session! The trainer explained complex concepts very clearly.',
        'date': '2025-06-20',
        'status': 'reviewed',
        'response': 'Thank you for your feedback! I\'m glad you found the session helpful.'
    }
    
    return render_template('trainer/feedback_detail.html', feedback=feedback_data)

@app.route('/trainer/message/<int:message_id>')
def trainer_message_detail(message_id):
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    # Mock message data - you can create a messages.json file later
    message_data = {
        'id': message_id,
        'from_name': 'Jane Smith',
        'from_email': 'jane.smith@example.com',
        'subject': 'Question about AI assignment',
        'content': 'Hi, I have a question about the AI assignment we discussed in class. Could you please clarify the requirements?',
        'date': '2025-06-20',
        'status': 'unread',
        'priority': 'medium'
    }
    
    return render_template('trainer/message_detail.html', message=message_data)

@app.route('/trainer/resource/<int:resource_id>')
def trainer_resource_detail(resource_id):
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    # Mock resource data - you can create a resources.json file later
    resource_data = {
        'id': resource_id,
        'title': 'Introduction to AI - Video Series',
        'type': 'video',
        'subject': 'ai',
        'description': 'Comprehensive video series covering AI fundamentals including machine learning, neural networks, and practical applications.',
        'duration': '2 hours',
        'difficulty': 'Beginner',
        'views': 156,
        'created': '2025-06-10',
        'file_size': '500MB',
        'downloads': 45,
        'rating': 4.5,
        'tags': ['AI', 'Machine Learning', 'Beginner', 'Video']
    }
    
    return render_template('trainer/resource_detail.html', resource=resource_data)

@app.route('/trainer/student/<int:student_id>')
def trainer_student_detail(student_id):
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as trainer to access this page.')
        return redirect(url_for('login'))
    
    # Mock student data - you can create a students.json file later
    student_data = {
        'id': student_id,
        'name': 'John Doe',
        'email': 'john.doe@example.com',
        'grade': '10th Grade',
        'school': 'Springfield High School',
        'city': 'Springfield',
        'joined_date': '2025-01-15',
        'total_classes': 8,
        'completed_classes': 6,
        'average_rating': 4.2,
        'progress_percentage': 75,
        'interests': ['AI', 'Mathematics', 'Computer Science'],
        'achievements': ['First Place - AI Competition', 'Perfect Attendance - 3 months'],
        'recent_activity': [
            {'date': '2025-06-20', 'action': 'Completed AI Fundamentals Quiz', 'score': '85%'},
            {'date': '2025-06-18', 'action': 'Submitted AI Project', 'status': 'Pending Review'},
            {'date': '2025-06-15', 'action': 'Attended AI Workshop', 'duration': '2 hours'}
        ]
    }
    
    return render_template('trainer/student_detail.html', student=student_data)

# General routes - Updated to render new templates
@app.route('/classes')
def classes():
    classes_data = get_collection_data('classes')
    return render_template('classes.html', classes=classes_data.get('classes', []))

@app.route('/resources')
def resources():
    return render_template('resources.html')

@app.route('/sessions')
def sessions():
    sessions_data = get_collection_data('sessions')
    classes_data = get_collection_data('classes')
    
    # Combine session data with class information
    combined_sessions = []
    for session in sessions_data.get('sessions', []):
        class_info = next((c for c in classes_data.get('classes', []) if c['id'] == session['class_id']), {})
        combined_session = {
            'id': session['id'],
            'title': class_info.get('title', 'Unknown Class'),
            'trainer': class_info.get('trainer', 'Unknown Trainer'),
            'date': class_info.get('date', 'Unknown Date'),
            'time': class_info.get('time', 'Unknown Time'),
            'duration': class_info.get('duration', 0),
            'status': 'completed',  # Default status
            'participants': 1,  # Default participant count
            'max_participants': class_info.get('max_participants', 0),
            'description': class_info.get('description', 'No description available'),
            'meeting_link': f"https://meet.google.com/session-{session['id']}"
        }
        combined_sessions.append(combined_session)
    
    return render_template('sessions.html', sessions=combined_sessions)

@app.route('/communications')
def communications():
    return render_template('communications.html')

@app.route('/community')
def community():
    return render_template('community.html')

@app.route('/reports')
def reports():
    return render_template('reports.html')

@app.route('/progress')
def progress():
    return render_template('progress.html')

@app.route('/skills')
def skills():
    return render_template('skills.html')

@app.route('/career')
def career():
    return render_template('career.html')

@app.route('/content')
def content():
    return render_template('content.html')

@app.route('/trainers')
def trainers():
    return render_template('trainers.html')

@app.route('/users')
def users():
    return render_template('users.html')

@app.route('/create-user')
def create_user():
    return render_template('create_user.html')

@app.route('/edit-user')
def edit_user():
    return render_template('edit_user.html')

@app.route('/achievements')
def achievements():
    return render_template('achievements.html')

@app.route('/doubts')
def doubts():
    return render_template('doubts.html')

@app.route('/interests')
def interests():
    return render_template('interests.html')

@app.route('/google-meet-settings')
def google_meet_settings():
    return render_template('google_meet_settings.html')

@app.route('/publish-training')
def publish_training():
    return render_template('publish_training.html')

@app.route('/attendance')
def attendance():
    return render_template('attendance.html')

# API Endpoints for dynamic data
@app.route('/api/classes')
def api_classes():
    classes_data = get_collection_data('classes')
    user_email = session.get('user_email')
    user_role = session.get('user_role')
    # Add 'enrolled' property for each class if student is logged in
    for cls in classes_data.get('classes', []):
        if user_role == 'student' and user_email:
            cls['enrolled'] = 'enrolled_students' in cls and user_email in cls['enrolled_students']
        else:
            cls['enrolled'] = False
    return jsonify(classes_data)

@app.route('/api/student/enrolled_classes')
def api_student_enrolled_classes():
    if 'user_email' not in session or session.get('user_role') != 'student':
        return jsonify({'error': 'Unauthorized'}), 401
    user_email = session['user_email']
    classes_data = get_collection_data('classes')
    enrolled = [cls for cls in classes_data.get('classes', []) if 'enrolled_students' in cls and user_email in cls['enrolled_students']]
    return jsonify({'enrolled_classes': enrolled})

@app.route('/api/student/available_classes')
def api_student_available_classes():
    if 'user_email' not in session or session.get('user_role') != 'student':
        return jsonify({'error': 'Unauthorized'}), 401
    user_email = session['user_email']
    classes_data = get_collection_data('classes')
    available = [cls for cls in classes_data.get('classes', []) if 'enrolled_students' not in cls or user_email not in cls['enrolled_students']]
    return jsonify({'available_classes': available})

@app.route('/api/sessions')
def api_sessions():
    sessions_data = get_collection_data('sessions')
    classes_data = get_collection_data('classes')
    
    # Combine session data with class information
    combined_sessions = []
    for session in sessions_data.get('sessions', []):
        class_info = next((c for c in classes_data.get('classes', []) if c['id'] == session['class_id']), {})
        combined_session = {
            'id': session['id'],
            'title': class_info.get('title', 'Unknown Class'),
            'trainer': class_info.get('trainer', 'Unknown Trainer'),
            'date': class_info.get('date', 'Unknown Date'),
            'time': class_info.get('time', 'Unknown Time'),
            'duration': class_info.get('duration', 0),
            'status': 'completed',  # Default status
            'participants': 1,  # Default participant count
            'max_participants': class_info.get('max_participants', 0),
            'description': class_info.get('description', 'No description available'),
            'meeting_link': f"https://meet.google.com/session-{session['id']}"
        }
        combined_sessions.append(combined_session)
    
    return jsonify({'sessions': combined_sessions})

@app.route('/api/resources')
def api_resources():
    # Mock resources data - you can create a resources.json file later
    resources = [
        {
            "id": 1,
            "title": "Introduction to AI - Video Series",
            "type": "video",
            "subject": "ai",
            "description": "Comprehensive video series covering AI fundamentals",
            "duration": "2 hours",
            "difficulty": "Beginner",
            "views": 156,
            "created": "2025-06-10"
        },
        {
            "id": 2,
            "title": "Mathematics Problem Solving Guide",
            "type": "document",
            "subject": "mathematics",
            "description": "PDF guide with practice problems and solutions",
            "pages": 45,
            "difficulty": "Intermediate",
            "views": 89,
            "created": "2025-06-12"
        },
        {
            "id": 3,
            "title": "Communication Skills Workshop",
            "type": "presentation",
            "subject": "communication",
            "description": "Interactive presentation on effective communication",
            "slides": 25,
            "difficulty": "Beginner",
            "views": 67,
            "created": "2025-06-15"
        }
    ]
    return jsonify({'resources': resources})

@app.route('/api/communications')
def api_communications():
    # Mock communications data
    communications = [
        {
            "id": 1,
            "type": "announcement",
            "title": "New AI Course Starting Next Week",
            "content": "We're excited to announce our new AI fundamentals course starting next week. Register now to secure your spot!",
            "author": "Admin Team",
            "date": "2025-06-15",
            "priority": "high"
        },
        {
            "id": 2,
            "type": "forum",
            "title": "Mathematics Problem Discussion",
            "content": "Let's discuss the quadratic equation problem from yesterday's session. Share your solutions!",
            "author": "Neha Gupta",
            "date": "2025-06-14",
            "replies": 5
        },
        {
            "id": 3,
            "type": "message",
            "title": "Assignment Submission Reminder",
            "content": "Don't forget to submit your communication skills assignment by Friday.",
            "author": "Anjali Desai",
            "date": "2025-06-13",
            "priority": "medium"
        }
    ]
    return jsonify({'communications': communications})

@app.route('/api/study-groups')
def api_study_groups():
    # Mock study groups data
    groups = [
        {
            "id": 1,
            "name": "Mathematics Enthusiasts",
            "subject": "Mathematics",
            "members": 45,
            "description": "Advanced problem solving and mathematical concepts"
        },
        {
            "id": 2,
            "name": "AI & ML Learners",
            "subject": "Artificial Intelligence",
            "members": 32,
            "description": "Exploring machine learning and AI fundamentals"
        },
        {
            "id": 3,
            "name": "Communication Skills",
            "subject": "Communication",
            "members": 28,
            "description": "Improving public speaking and communication"
        }
    ]
    return jsonify({'groups': groups})

@app.route('/api/discussions')
def api_discussions():
    # Mock discussions data
    discussions = [
        {
            "id": 1,
            "title": "Best study techniques for mathematics?",
            "author": "Priya Sharma",
            "replies": 12,
            "views": 89,
            "date": "2025-06-15"
        },
        {
            "id": 2,
            "title": "AI career opportunities discussion",
            "author": "Rahul Kumar",
            "replies": 8,
            "views": 67,
            "date": "2025-06-14"
        },
        {
            "id": 3,
            "title": "Communication tips for presentations",
            "author": "Anjali Desai",
            "replies": 15,
            "views": 123,
            "date": "2025-06-13"
        }
    ]
    return jsonify({'discussions': discussions})

@app.route('/api/community-feed')
def api_community_feed():
    # Mock community feed data
    posts = [
        {
            "id": 1,
            "type": "achievement",
            "author": "Neha Gupta",
            "content": "Just completed the AI fundamentals course! 🎉",
            "likes": 24,
            "comments": 8,
            "date": "2025-06-15"
        },
        {
            "id": 2,
            "type": "question",
            "author": "Rajesh Patel",
            "content": "Can anyone help me understand quadratic equations?",
            "likes": 5,
            "comments": 12,
            "date": "2025-06-14"
        },
        {
            "id": 3,
            "type": "resource",
            "author": "Dr. Priya Sharma",
            "content": "Shared a new study guide for mathematics. Check it out!",
            "likes": 18,
            "comments": 6,
            "date": "2025-06-13"
        }
    ]
    return jsonify({'posts': posts})

@app.route('/api/recent-reports')
def api_recent_reports():
    # Mock recent reports data
    reports = [
        {
            "id": 1,
            "name": "Student Performance Report - June 2025",
            "type": "performance",
            "generated": "2025-06-15",
            "status": "completed",
            "size": "2.3 MB"
        },
        {
            "id": 2,
            "name": "Course Analytics - AI Fundamentals",
            "type": "course",
            "generated": "2025-06-14",
            "status": "completed",
            "size": "1.8 MB"
        },
        {
            "id": 3,
            "name": "Attendance Report - Week 24",
            "type": "attendance",
            "generated": "2025-06-13",
            "status": "completed",
            "size": "0.9 MB"
        }
    ]
    return jsonify({'reports': reports})

@app.route('/api/course-progress')
def api_course_progress():
    # Mock course progress data
    courses = [
        {
            "id": 1,
            "title": "Introduction to AI",
            "progress": 85,
            "completed": 17,
            "total": 20,
            "lastAccessed": "2025-06-15"
        },
        {
            "id": 2,
            "title": "Advanced Mathematics",
            "progress": 60,
            "completed": 12,
            "total": 20,
            "lastAccessed": "2025-06-14"
        },
        {
            "id": 3,
            "title": "Communication Skills",
            "progress": 100,
            "completed": 15,
            "total": 15,
            "lastAccessed": "2025-06-13"
        }
    ]
    return jsonify({'courses': courses})

@app.route('/api/learning-path')
def api_learning_path():
    # Mock learning path data
    path = [
        {
            "id": 1,
            "title": "Foundation Skills",
            "status": "completed",
            "description": "Basic programming and mathematics"
        },
        {
            "id": 2,
            "title": "AI Fundamentals",
            "status": "in-progress",
            "description": "Introduction to artificial intelligence"
        },
        {
            "id": 3,
            "title": "Advanced AI",
            "status": "locked",
            "description": "Deep learning and neural networks"
        },
        {
            "id": 4,
            "title": "AI Applications",
            "status": "locked",
            "description": "Real-world AI implementations"
        }
    ]
    return jsonify({'path': path})

@app.route('/api/achievements')
def api_achievements():
    # Mock achievements data
    achievements = [
        {
            "id": 1,
            "title": "First Course Completed",
            "description": "Successfully completed your first course",
            "date": "2025-06-15",
            "icon": "fa-trophy"
        },
        {
            "id": 2,
            "title": "Perfect Score",
            "description": "Achieved 100% on a quiz",
            "date": "2025-06-14",
            "icon": "fa-star"
        },
        {
            "id": 3,
            "title": "Study Streak",
            "description": "Studied for 7 consecutive days",
            "date": "2025-06-13",
            "icon": "fa-fire"
        }
    ]
    return jsonify({'achievements': achievements})

@app.route('/api/my-skills')
def api_my_skills():
    # Mock skills data
    skills = [
        {
            "id": 1,
            "name": "Python Programming",
            "category": "technical",
            "level": 75,
            "description": "Intermediate level Python programming skills"
        },
        {
            "id": 2,
            "name": "Public Speaking",
            "category": "communication",
            "level": 60,
            "description": "Basic public speaking and presentation skills"
        },
        {
            "id": 3,
            "name": "Problem Solving",
            "category": "creativity",
            "level": 80,
            "description": "Strong analytical and problem-solving abilities"
        },
        {
            "id": 4,
            "name": "Team Leadership",
            "category": "leadership",
            "level": 45,
            "description": "Basic team management and leadership skills"
        }
    ]
    return jsonify({'skills': skills})

@app.route('/api/recommended-skills')
def api_recommended_skills():
    # Mock recommended skills data
    skills = [
        {
            "id": 5,
            "name": "Machine Learning",
            "category": "technical",
            "priority": "high",
            "description": "Learn ML algorithms and frameworks",
            "estimatedTime": "3 months"
        },
        {
            "id": 6,
            "name": "Data Analysis",
            "category": "technical",
            "priority": "medium",
            "description": "Master data analysis and visualization",
            "estimatedTime": "2 months"
        },
        {
            "id": 7,
            "name": "Conflict Resolution",
            "category": "leadership",
            "priority": "medium",
            "description": "Develop conflict resolution skills",
            "estimatedTime": "1 month"
        }
    ]
    return jsonify({'skills': skills})

@app.route('/api/development-path')
def api_development_path():
    # Mock development path data
    path = [
        {
            "id": 1,
            "title": "Foundation Phase",
            "skills": ["Python Programming", "Basic Mathematics"],
            "status": "completed",
            "duration": "2 months"
        },
        {
            "id": 2,
            "title": "Intermediate Phase",
            "skills": ["Machine Learning", "Data Analysis"],
            "status": "in-progress",
            "duration": "3 months"
        },
        {
            "id": 3,
            "title": "Advanced Phase",
            "skills": ["Deep Learning", "AI Applications"],
            "status": "locked",
            "duration": "4 months"
        }
    ]
    return jsonify({'path': path})

# API endpoints for interactive features
@app.route('/api/study-groups/<int:group_id>/join', methods=['POST'])
def api_join_study_group(group_id):
    if 'user_email' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    # Mock implementation - in real app, you'd update the database
    return jsonify({'success': True, 'message': 'Successfully joined the study group'})

@app.route('/api/posts/<int:post_id>/like', methods=['POST'])
def api_like_post(post_id):
    if 'user_email' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    # Mock implementation - in real app, you'd update the database
    return jsonify({'success': True, 'message': 'Post liked successfully'})

@app.route('/api/posts/<int:post_id>/comment', methods=['POST'])
def api_comment_post(post_id):
    if 'user_email' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    data = request.get_json()
    comment = data.get('comment', '')
    
    if not comment:
        return jsonify({'error': 'Comment cannot be empty'}), 400
    
    # Mock implementation - in real app, you'd save to database
    return jsonify({'success': True, 'message': 'Comment posted successfully'})

@app.route('/session/<int:session_id>/join')
def join_session(session_id):
    if 'user_email' not in session:
        flash('Please login to join sessions.', 'warning')
        return redirect(url_for('login'))
    
    # Mock implementation - in real app, you'd redirect to actual meeting
    flash(f'Joining session {session_id}...', 'info')
    return redirect(url_for('sessions'))

@app.route('/resource/<int:resource_id>')
def view_resource(resource_id):
    if 'user_email' not in session:
        flash('Please login to access resources.', 'warning')
        return redirect(url_for('login'))
    
    # Mock implementation - in real app, you'd serve the actual resource
    flash(f'Accessing resource {resource_id}...', 'info')
    return redirect(url_for('resources'))

@app.route('/communication/<int:comm_id>')
def view_communication(comm_id):
    # Mock implementation
    flash(f'Viewing communication {comm_id}...', 'info')
    return redirect(url_for('communications'))

@app.route('/discussion/<int:discussion_id>')
def view_discussion(discussion_id):
    # Mock implementation
    flash(f'Viewing discussion {discussion_id}...', 'info')
    return redirect(url_for('community'))

@app.route('/report/<int:report_id>')
def view_report(report_id):
    # Mock implementation
    flash(f'Viewing report {report_id}...', 'info')
    return redirect(url_for('reports'))

# API endpoints for dynamic functionality
@app.route('/api/enroll/<int:class_id>', methods=['POST'])
def api_enroll_class(class_id):
    """Enroll a student in a class"""
    if 'user_email' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    if session['user_role'] != 'student':
        return jsonify({'error': 'Only students can enroll in classes'}), 403
    
    try:
        # Load existing data
        classes_data = get_collection_data('classes')
        users_data = get_collection_data('users')
        
        # Find the class
        class_found = None
        for cls in classes_data.get('classes', []):
            if cls['id'] == class_id:
                class_found = cls
                break
        
        if not class_found:
            return jsonify({'error': 'Class not found'}), 404
        
        # Check if already enrolled
        user_email = session['user_email']
        if 'enrolled_students' not in class_found:
            class_found['enrolled_students'] = []
        
        if user_email in class_found['enrolled_students']:
            return jsonify({'error': 'Already enrolled in this class'}), 400
        
        # Enroll the student
        class_found['enrolled_students'].append(user_email)
        
        # Update user's enrolled classes
        if user_email in users_data:
            if 'enrolled_classes' not in users_data[user_email]:
                users_data[user_email]['enrolled_classes'] = []
            users_data[user_email]['enrolled_classes'].append(class_id)
        
        # Save updated data
        set_collection_data('classes', classes_data)
        set_collection_data('users', users_data)
        
        return jsonify({
            'success': True,
            'message': f'Successfully enrolled in {class_found["title"]}',
            'class_id': class_id
        })
        
    except Exception as e:
        print(f"Error enrolling in class: {e}")
        return jsonify({'error': 'Error enrolling in class. Please try again.'}), 500

@app.route('/api/unenroll/<int:class_id>', methods=['POST'])
def api_unenroll_class(class_id):
    """Unenroll a student from a class"""
    if 'user_email' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    if session['user_role'] != 'student':
        return jsonify({'error': 'Only students can unenroll from classes'}), 403
    
    try:
        # Load existing data
        classes_data = get_collection_data('classes')
        users_data = get_collection_data('users')
        
        # Find the class
        class_found = None
        for cls in classes_data.get('classes', []):
            if cls['id'] == class_id:
                class_found = cls
                break
        
        if not class_found:
            return jsonify({'error': 'Class not found'}), 404
        
        # Check if enrolled
        user_email = session['user_email']
        if 'enrolled_students' not in class_found:
            class_found['enrolled_students'] = []
        
        if user_email not in class_found['enrolled_students']:
            return jsonify({'error': 'Not enrolled in this class'}), 400
        
        # Unenroll the student
        class_found['enrolled_students'].remove(user_email)
        
        # Update user's enrolled classes
        if user_email in users_data and 'enrolled_classes' in users_data[user_email]:
            if class_id in users_data[user_email]['enrolled_classes']:
                users_data[user_email]['enrolled_classes'].remove(class_id)
        
        # Save updated data
        set_collection_data('classes', classes_data)
        set_collection_data('users', users_data)
        
        return jsonify({
            'success': True,
            'message': f'Successfully unenrolled from {class_found["title"]}',
            'class_id': class_id
        })
        
    except Exception as e:
        print(f"Error unenrolling from class: {e}")
        return jsonify({'error': 'Error unenrolling from class. Please try again.'}), 500

# Individual class view route
@app.route('/class/<int:class_id>')
def view_class(class_id):
    """View individual class details"""
    classes_data = get_collection_data('classes')
    
    # Find the class
    class_found = None
    for cls in classes_data.get('classes', []):
        if cls['id'] == class_id:
            class_found = cls
            break
    
    if not class_found:
        flash('Class not found.', 'danger')
        return redirect(url_for('classes'))
    
    # Check if user is enrolled (if logged in)
    is_enrolled = False
    if 'user_email' in session:
        user_email = session['user_email']
        if 'enrolled_students' in class_found and user_email in class_found['enrolled_students']:
            is_enrolled = True
    
    return render_template('class_detail.html', class_data=class_found, is_enrolled=is_enrolled)

# Trainer reports route
@app.route('/trainer/reports')
def trainer_reports():
    if 'user_role' not in session or session['user_role'] != 'trainer':
        flash('Please login as a trainer to access this page.')
        return redirect(url_for('login'))
    
    return render_template('trainer/reports.html')

@app.route('/api/classes/add', methods=['POST'])
def api_add_class():
    if session.get('user_role') != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    data = request.get_json()
    required = ['name', 'description', 'start_date', 'end_date', 'trainer_email']
    if not all(k in data and data[k] for k in required):
        return jsonify({'success': False, 'error': 'Missing fields'}), 400
    # Load existing classes
    try:
        with open('classes.json', 'r') as f:
            classes = json.load(f)
    except Exception:
        classes = []
    # Assign a new id
    new_id = max([c.get('id', 0) for c in classes] or [0]) + 1
    new_class = {
        'id': new_id,
        'name': data['name'],
        'description': data['description'],
        'start_date': data['start_date'],
        'end_date': data['end_date'],
        'trainer_email': data['trainer_email']
    }
    classes.append(new_class)
    try:
        with open('classes.json', 'w') as f:
            json.dump(classes, f, indent=2)
    except Exception as e:
        return jsonify({'success': False, 'error': 'Failed to save class.'}), 500
    return jsonify({'success': True, 'class': new_class})

@app.route('/student/career-assessment', methods=['GET', 'POST'])
def student_career_assessment():
    if request.method == 'POST':
        name = request.form.get('name')
        age = request.form.get('age')
        email = request.form.get('email')
        q1 = request.form.get('q1')
        q2 = request.form.get('q2')
        q3 = request.form.get('q3')
        q4 = request.form.get('q4')
        q5 = request.form.get('q5')
        
        # Save to Excel
        file_path = 'career_assessments.xlsx'
        try:
            from openpyxl import load_workbook, Workbook
            import os
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(['Timestamp', 'Name', 'Age', 'Email', 'Q1', 'Q2', 'Q3', 'Q4', 'Q5'])
            ws.append([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                name, age, email, q1, q2, q3, q4, q5
            ])
            wb.save(file_path)
        except Exception as e:
            flash(f'Error saving assessment: {e}', 'danger')
            return redirect(url_for('student_career_assessment'))
        flash('Your assessment has been submitted successfully!', 'success')
        return redirect(url_for('student_career_assessment'))
    return render_template('student/career.html')

# Add new route for profile completion
@app.route('/student/complete_profile', methods=['GET', 'POST'])
def complete_student_profile():
    if 'user_email' not in session or session.get('user_role') != 'student':
        flash('Please login as a student to complete your profile.')
        return redirect(url_for('login'))
    users_data = get_collection_data('users')
    user = users_data.get(session['user_email'], {})
    if request.method == 'POST':
        grade = request.form.get('grade')
        dob = request.form.get('dob')
        gender = request.form.get('gender')
        school = request.form.get('school')
        city = request.form.get('city')
        bio = request.form.get('bio', '')
        subjects = request.form.getlist('interested_subjects')
        # Validate required fields
        if not (grade and dob and gender and school and city and subjects):
            flash('Please fill in all required fields and select at least one subject.')
            return render_template('student/complete_profile.html', user=user)
        user['grade'] = grade
        user['dob'] = dob
        user['gender'] = gender
        user['school'] = school
        user['city'] = city
        user['bio'] = bio
        user['interested_subjects'] = subjects
        users_data[session['user_email']] = user
        set_collection_data('users', users_data)
        flash('Profile completed successfully!')
        return redirect(url_for('student_dashboard'))
    return render_template('student/complete_profile.html', user=user)

@app.route('/student/class/<int:class_id>')
def student_view_class(class_id):
    # Optionally, you can add student-specific logic here
    return view_class(class_id)

# MathStar 3D Game Route
@app.route('/mathstar-3d')
def mathstar_3d():
    """MathStar 3D - Interactive Math Learning Game for Kids"""
    return render_template('mathstar_3d.html')

@app.route('/english-star')
def english_star():
    """English Star - Interactive English Learning Game"""
    return render_template('english_star.html')

@app.route('/science-star')
def science_star():
    """Science Star - Interactive Science Learning Game"""
    return render_template('science_star.html')

@app.route('/history-star')
def history_star():
    """History Star - Interactive History Learning Game"""
    return render_template('history_star.html')

@app.route('/computer-star')
def computer_star():
    """Computer Star - Interactive Computer Science Learning Game"""
    return render_template('computer_star.html')

@app.route('/ai-star')
def ai_star():
    """AI Star - Interactive AI Learning Game"""
    return render_template('ai_star.html')

@app.route('/geography-star')
def geography_star():
    """Geography Star - Interactive Geography Learning Game"""
    return render_template('geography_star.html')

# Game Hub Route - Central navigation for all games
@app.route('/game-hub')
def game_hub():
    """Game Hub - Central navigation for all educational games"""
    return render_template('game_hub.html')

if __name__ == '__main__':
    # Add ProxyFix middleware
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
    
    app.run(debug=True, host='0.0.0.0', port=8000)
